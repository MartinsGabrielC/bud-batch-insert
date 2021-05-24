# encoding: utf-8
from Model.KeyResult import KeyResult
from Model.Objective import Objective
from Model.Database import *
from Model.Team import Team
from Model.User import User
from Model.UserTeam import UserTeam
from Model.Cycle import Cycle
from Util.Constants import *
from Util.PasswordUtil import get_strong_password
from auth0.v3.management import Auth0
from auth0.v3.authentication import GetToken
from auth0.v3.exceptions import Auth0Error
from urllib3.exceptions import NewConnectionError
from tqdm import tqdm
import psycopg2
from psycopg2.errors import UniqueViolation
import openpyxl
import logging
import boto3
import threading


def connect(config):
    """ 
    Connect to the PostgreSQL database server and return connection
    -------
    Parameters
    -------
    config: dict
        Contains environment configuration
    -------
    Returns
    -------
    conn
        PostgreSQL connection
    
    """

    host = config["POSTGRES_HOST"]
    port = config["POSTGRES_PORT"]
    database = config["POSTGRES_DB"]
    user = config["POSTGRES_USER"]
    try:
        password = config["POSTGRES_PASSWORD"]
    except KeyError:
        pass

    try:
        region = config["AWS_REGION"]
        client = boto3.client('rds',
                                aws_access_key_id=config["AWS_CREDENTIALS_ACCESS_KEY_ID"],
                                aws_secret_access_key=config["AWS_CREDENTIALS_SECRET_ACCESS_KEY"])
        password = client.generate_db_auth_token(DBHostname=host, Port=port, DBUsername=user, Region=region)
    except KeyError:
        pass

    conn = None
    try:
        # connect to the PostgreSQL server
        logging.debug("Util : Creating conection to PostgreSQL")
        conn = psycopg2.connect(
            host=host,
            port=port,
            database=database,
            user=user,
            password=password
        )
        conn.set_session(autocommit=True)
        return conn
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)

def parse_user(ws):
    users = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=9):
        #Check for Mandatory Fields
        result = True
        for position in USER_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            user = User.fromlist(values)
            users.append(user)
    return users

def parse_team(ws):
    teams = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=5):
        #Check for Mandatory Fields
        result = True
        for position in TEAM_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            team = Team.fromlist(values)
            teams.append(team)
    return teams

def parse_user_team(ws):
    userteams = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=2):
        #Check for Mandatory Fields
        result = True
        for position in USER_X_TEAM_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            userteam = UserTeam.fromlist(values)
            userteams.append(userteam)
    return userteams

def parse_cycle(ws):
    cycles = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=6):
        #Check for Mandatory Fields
        result = True
        for position in CYCLE_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            cycle = Cycle.fromlist(values)
            cycles.append(cycle)
    return cycles

def parse_objective(ws):
    objectives = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=6):
        #Check for Mandatory Fields
        result = True
        for position in OBJECTIVE_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            objective = Objective.fromlist(values)
            objectives.append(objective)
    return objectives

def parse_key_result(ws):
    key_results = []
    for row in ws.iter_rows(min_row=2,  min_col=0, max_col=10):
        #Check for Mandatory Fields
        result = True
        for position in KEY_RESULTS_MANDATORY:
            result = result and (row[position].value != None and row[position].value != "")
        if result:
            values = [item.value if item.value != None else "" for item in row] #Convert None to empty strings
            key_result = KeyResult.fromlist(values)
            key_results.append(key_result)
    return key_results

def readSheet(wb, sheet, results):
    if sheet in VALID_TABS:
        ws = wb[sheet]
        if sheet == USER_TAB:
            results[USER_TAB] += parse_user(ws)
        if sheet == TEAM_TAB:
            results[TEAM_TAB] += parse_team(ws)
        if sheet == USER_X_TEAM_TAB:
            results[USER_X_TEAM_TAB] += parse_user_team(ws)
        if sheet == CYCLE_TAB:
            results[CYCLE_TAB] += parse_cycle(ws)
        if sheet == OBJECTIVE_TAB:
            results[OBJECTIVE_TAB] += parse_objective(ws)
        if sheet == KEY_RESULTS_TAB:
            results[KEY_RESULTS_TAB] += parse_key_result(ws)

def read_excel(config, path_excel, file_list, results):
    for file in file_list:
        logging.debug("Util : Reading xlsx file - \"{}\"".format(file))
        threads = list()
        if file.endswith(".xlsx"):
            wb = openpyxl.load_workbook(path_excel+file)
            for index, sheet in enumerate(VALID_TABS):
                #Each sheet runs in a different thread
                logging.debug("      Read Excel : create and start thread %d - %s.", index, sheet)
                x = threading.Thread(target=readSheet, args=(wb, sheet, results))
                threads.append(x)
                x.start()
            for index, thread in enumerate(threads):
                thread.join()
                logging.debug("      Read Excel : thread %d done", index)
            wb.save(path_excel+file)

def load_user_to_authz(config, user):
    domain = config["AUTHZ_DOMAIN"]
    non_interactive_client_id = config["AUTHZ_CLIENT_ID"]
    non_interactive_client_secret = config["AUTHZ_CLIENT_SECRET"]
    get_token = GetToken(domain)
    try:
        token = get_token.client_credentials(non_interactive_client_id,
            non_interactive_client_secret, 'https://{}/api/v2/'.format(domain))
        mgmt_api_token = token['access_token']
    except ConnectionError:
        raise Exception("AUTH0: Error - Host is unavailable")
    auth0 = Auth0(domain, mgmt_api_token)
    secure_password = get_strong_password()

    name = user.get_first_name() + user.get_last_name()

    body = {
    "email": user.get_email(),
    "name": name,
    "connection": "Username-Password-Authentication",
    "password": secure_password,
    }
    if user.get_picture() != "":
        body["picture"] =  user.get_picture()
    try:
        auth0user = auth0.users.create(body)
        return auth0user["user_id"]
    except Auth0Error:
        f = open(config["LOGFILE"], "a")
        f.write("AUTH0: User {} already exists\n".format(user.get_email()))
        f.close()
        return ""
    except (ConnectionError, NewConnectionError):
        raise Exception("AUTH0: Error - Host is unavailable")

def send_confirmation_email_authz(config, user):
    domain = config["AUTHZ_DOMAIN"]
    non_interactive_client_id = config["AUTHZ_CLIENT_ID"]
    non_interactive_client_secret = config["AUTHZ_CLIENT_SECRET"]
    get_token = GetToken(domain)
    try:
        token = get_token.client_credentials(non_interactive_client_id,
            non_interactive_client_secret, 'https://{}/api/v2/'.format(domain))
        mgmt_api_token = token['access_token']
    except ConnectionError:
        raise Exception("AUTH0: Error - Host is unavailable")
    auth0 = Auth0(domain, mgmt_api_token)
    secure_password = get_strong_password()

    conn = connect(config)

    sql = """ select authz_sub from public."user" where email=%s limit 1;"""
    cur = conn.cursor()
    email = user.get_email()
    print(email)
    cur.execute(sql, (email, ))
    result = cur.fetchone()
    print(result[0])
    body = {
    "user_id": result[0]
    }
    cur.close()
    conn.close()
    #auth0.jobs.send_verification_email(body)
    
def load_users_to_database(config, users):
    conn = connect(config)
    sql = """ INSERT INTO public."user"(first_name, last_name, nickname, linked_in_profile_address, role, email, picture, gender, authz_sub)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id;"""
    cur = conn.cursor()
    for user in users:
        userid = load_user_to_authz(config, user)
        if userid != "":
            user.set_authz_sub(userid)
            cur.execute(sql, user.get_values_tuple())
    cur.close()
    conn.close()

def load_teams_to_database(config, teams):
    conn = connect(config)
    sql = """ INSERT INTO public."team"(name, description, parent_id, gender, owner_id)
            SELECT %s,%s, (SELECT id from team where name=%s),%s, (SELECT id from "user" where email=%s)
            WHERE NOT EXISTS ( SELECT id FROM public."team" WHERE name=%s );"""
    cur = conn.cursor()
    for team in teams:
        values = list(team.get_values_tuple())
        values.append(team.get_name())
        cur.execute(sql, tuple(values))
    cur.close()
    conn.close()

def load_user_teams_to_database(config, userteams):
    conn = connect(config)
    sql = """ INSERT INTO public."team_users_user"(user_id, team_id)
            SELECT (SELECT id from public."user" where email=%s), (SELECT id from public."team" where name=%s)
            WHERE NOT EXISTS ( SELECT user_id, team_id FROM public."team_users_user" WHERE user_id=(SELECT id from public."user" where email=%s) 
                                                                                        AND team_id=(SELECT id from public."team" where name=%s));"""
    cur = conn.cursor()
    for userteam in userteams:
        values = list(userteam.get_values_tuple())
        values.extend([userteam.get_user_email(),userteam.get_team_name()])
        cur.execute(sql, tuple(values))
    cur.close()
    conn.close()

def load_cycle_to_database(config, cycles):
    conn = connect(config)
    sql = """ INSERT INTO public."cycle"(period, cadence, date_start, date_end, team_id, parent_id)
            SELECT %s,%s,%s,%s,(SELECT id from public."team" where name=%s), (SELECT id from public."cycle" where period=%s)
            WHERE NOT EXISTS ( SELECT id FROM public."cycle" WHERE period=%s AND team_id=(SELECT id from public."team" where name=%s));"""
    cur = conn.cursor()
    for cycle in cycles:
        values = list(cycle.get_values_tuple())
        values.extend([cycle.get_period(),cycle.get_team_name()])
        cur.execute(sql, tuple(values))
    cur.close()
    conn.close()

def load_objective_to_database(config, objectives):
    conn = connect(config)
    sql = """ INSERT INTO public."objective"(title, cycle_id, owner_id)
            SELECT %s,(SELECT id FROM public."cycle" WHERE period=%s AND team_id=(SELECT id FROM public."team" WHERE name=%s)),(SELECT id from public."user" where email=%s) 
            WHERE NOT EXISTS ( SELECT id FROM public."objective" WHERE title=%s AND cycle_id=(SELECT id FROM public."cycle" WHERE period=%s AND team_id=(SELECT id FROM public."team" WHERE name=%s)));"""
    cur = conn.cursor()
    for objective in objectives:
        values = list(objective.get_values_tuple())
        values.extend([objective.get_title(),objective.get_cycle_period(),objective.get_team_name()])
        cur.execute(sql, tuple(values))
    cur.close()
    conn.close()

def load_key_result_to_database(config, key_results):
    conn = connect(config)
    sql = """ INSERT INTO public."key_result"(title, description, type, format, goal, initial_value, objective_id, owner_id, team_id)
            SELECT %s,%s,%s,%s,%s,%s,(SELECT id FROM public."objective" WHERE cycle_id=(SELECT id FROM public."cycle" WHERE period=%s AND team_id=(SELECT id FROM public."team" WHERE name=%s)) AND title=%s), (SELECT id from public."user" where email=%s), (SELECT id from public."team" where name=%s)
            WHERE NOT EXISTS ( SELECT id FROM public."key_result" WHERE title=%s AND objective_id=(SELECT id FROM public."objective" WHERE cycle_id=(SELECT id FROM public."cycle" WHERE period=%s AND team_id=(SELECT id FROM public."team" WHERE name=%s)) AND title=%s));"""
    cur = conn.cursor()
    for key_result in key_results:
        values = list(key_result.get_values_tuple())
        values.extend([key_result.get_team_name(), key_result.get_title(), key_result.get_cycle_period(), key_result.get_team_name(), key_result.get_objective_title()])
        cur.execute(sql, tuple(values))
    cur.close()
    conn.close()

def load_to_database(config, data):
    load_users_to_database(config, tqdm(data[USER_TAB], desc="USERS", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))
    load_teams_to_database(config, tqdm(data[TEAM_TAB], desc="TEAMS", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))
    load_user_teams_to_database(config, tqdm(data[USER_X_TEAM_TAB], desc="USERS X TEAM", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))
    load_cycle_to_database(config, tqdm(data[CYCLE_TAB], desc="CYCLES", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))
    load_objective_to_database(config, tqdm(data[OBJECTIVE_TAB], desc="OBJECTIVES", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))
    load_key_result_to_database(config, tqdm(data[KEY_RESULTS_TAB], desc="KEY RESULTS", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"))

def verificate_users(config, users):
    for user in tqdm(users, desc="USERS", bar_format="{desc:<15}{percentage:3.0f}%|{bar}{r_bar}"):
        send_confirmation_email_authz(config, user)